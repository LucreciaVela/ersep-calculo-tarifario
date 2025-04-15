
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import shutil
from io import BytesIO

st.set_page_config(page_title="ERSEP Transporte", layout="centered")
st.image("ersep_logo.png", width=200)
st.title("🚌 Calculadora Tarifaria - ERSEP Transporte")
st.markdown("Ingresá los valores necesarios y presioná **Calcular Tarifa** para obtener el resultado tarifario.")

items_visibles = {
    "MT": "Monto anual de la prima para personal de conducción",
    "U": "Costo de los uniformes según convenio. Por temporada: 1 campera + 2 camisas + 1 pantalón + 2 corbatas",
    "Nc": "Unidad de valuación (medidas 900 R22,5) + 0,6 (1 cub. recap.)",
    "Nm": "Unidad de valuación (medidas 275/80 R22,5) + 0,6 (1 cub. recap.)",
    "Ng": "Unidad de valuación (medidas 295/80 R22,5) + 0,6 (1 cub. recap.)",
    "L": "Precio de aceite para motor (tambor de 205 litros)",
    "Mp": "Monto de la prima mensual en pesos",
    "E": "Costo de lavado y engrase",
    "Pp": "Monto anual de patente ponderado",
    "RTM": "Recorrido Total Mensual",
    "Sbcu": "Sueldo básico de conducción - empresas metropolitanas",
    "Pm": "Porcentaje de participación empresas metropolitanas",
    "Pr": "Porcentaje de participación empresas rurales",
    "SBcg": "Sueldo básico de conducción - empresas rurales",
    "Gm": "Precio Gas Oil para empresas metropolitanas",
    "Gr": "Precio Gas Oil para empresas rurales",
    "Vc": "Valor unidad nueva chica",
    "Vm": "Valor unidad nueva mediana",
    "Vg": "Valor unidad nueva grande",
    "RBM": "Recaudación Bruta Mensual Promedio",
    "Ut": "Flota Total Pcia de Córdoba - Cantidad de Vehículos"
}

@st.cache_data
def cargar_referencia():
    hoja = pd.read_excel("Incremento TBK_ Mesa 13 Octubre 2025.xlsx", sheet_name="Hoja Llave", header=None)
    nombres = hoja.iloc[:23, 0].tolist()
    valores_ref = hoja.iloc[:23, 14].tolist()
    ref_dict = {str(nombres[i]): valores_ref[i] for i in range(len(nombres)) if str(nombres[i]) in items_visibles}
    return ref_dict

def actualizar_excel_con_datos(entradas_usuario):
    ruta_original = "Incremento TBK_ Mesa 13 Octubre 2025.xlsx"
    ruta_temp = "ersep_calculo_actualizado.xlsx"
    shutil.copy(ruta_original, ruta_temp)

    wb = load_workbook(ruta_temp, data_only=False)
    hoja = wb["Hoja Llave"]

    def es_combinada(hoja, fila, columna):
        for rango in hoja.merged_cells.ranges:
            if (fila, columna) in rango.cells:
                return True
        return False

    fila_inicio = 3
    fila_actual = fila_inicio
    i = 0
    claves = list(items_visibles.keys())
    while i < len(claves) and fila_actual < 200:
        clave = hoja.cell(row=fila_actual, column=1).value
        if clave in claves and not es_combinada(hoja, fila_actual, 2):
            hoja.cell(row=fila_actual, column=2).value = entradas_usuario[i]
            i += 1
        fila_actual += 1

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida, load_workbook(salida, data_only=True)

def obtener_resumen(hoja):
    datos = []
    for fila in hoja.iter_rows(min_row=4, max_col=10, max_row=100):
        c1 = fila[0].value
        c2 = fila[1].value
        c3 = fila[9].value if len(fila) > 9 else None
        if c1 or c2 or c3:
            datos.append((c1, c2, c3))
    return datos

referencias = cargar_referencia()
entradas_usuario = []

with st.form("formulario_datos"):
    st.markdown("### Ingreso de datos")
    for clave, descripcion in items_visibles.items():
        valor_ref = referencias.get(clave, 0.0)
        st.markdown(f"**🔹 {clave}**  
{descripcion}")
        entrada = st.number_input(label="", value=float(valor_ref) if isinstance(valor_ref, (int, float)) else 0.0, step=1.0, key=f"input_{clave}")
        entradas_usuario.append(entrada)
    calcular = st.form_submit_button("Calcular Tarifa")

if calcular:
    archivo_excel, wb_final = actualizar_excel_con_datos(entradas_usuario)
    hoja_resumen = wb_final["Resumen de Calculo"]
    datos = obtener_resumen(hoja_resumen)

    st.success("✅ Cálculo realizado. A continuación se muestra la hoja resumen.")
    st.write("### 📄 Hoja: Resumen de Cálculo")

    df_resumen = pd.DataFrame([{"Código": f[0], "Concepto": f[1], "Valor": f[2]} for f in datos if f[1]])
    df_resumen.dropna(subset=["Valor"], inplace=True)

    total = df_resumen["Valor"].sum()
    df_resumen["% Incidencia"] = df_resumen["Valor"] / total * 100

    st.dataframe(df_resumen.style.format({"Valor": "{:,.2f}", "% Incidencia": "{:.2f}%"}), use_container_width=True)

    resumen_salida = BytesIO()
    with pd.ExcelWriter(resumen_salida, engine="xlsxwriter") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen de Cálculo", index=False)
    resumen_salida.seek(0)

    st.download_button("📥 Descargar Hoja Resumen en Excel", data=resumen_salida, file_name="Resumen_Calculo_ERSEP.xlsx")
