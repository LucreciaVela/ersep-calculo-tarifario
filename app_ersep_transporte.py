
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import shutil
from io import BytesIO

st.set_page_config(page_title="ERSEP Transporte", layout="centered")
st.title("🚌 Calculadora Tarifaria - ERSEP Transporte")
st.markdown("Ingresá los valores necesarios y presioná **Calcular Tarifa** para obtener el resultado tarifario.")

# Ítems visibles definidos por el usuario
nombres_visibles = [
    "MT", "U", "Nc", "Nm", "Ng", "L", "Mp", "E", "Pp", "RTM",
    "Sbcu", "Pm", "Pr", "SBcg", "Gm", "Gr", "Vc", "Vm", "Vg", "RBM", "Ut"
]

@st.cache_data
def cargar_referencia():
    hoja = pd.read_excel("Incremento TBK_ Mesa 13 Octubre 2025.xlsx", sheet_name="Hoja Llave", header=None)
    nombres = hoja.iloc[:23, 0].tolist()
    valores_ref = hoja.iloc[:23, 14].tolist()
    etiquetas = {str(n): str(nombres[i]) for i, n in enumerate(nombres) if pd.notna(n)}
    return etiquetas, valores_ref

def actualizar_excel_con_datos(entradas_usuario, etiquetas):
    ruta_original = "Incremento TBK_ Mesa 13 Octubre 2025.xlsx"
    ruta_temp = "ersep_calculo_actualizado.xlsx"
    shutil.copy(ruta_original, ruta_temp)

    wb = load_workbook(ruta_temp, data_only=False)
    hoja = wb["Hoja Llave"]

    def es_combinada(hoja, fila, columna):
        for rango in hoja.merged_cells.ranges:
            if (fila, columna) in rango.cells:
                return True
        return False

    fila_inicio = 3
    fila_actual = fila_inicio
    i = 0
    while i < len(nombres_visibles) and fila_actual < 200:
        clave = hoja.cell(row=fila_actual, column=1).value
        if clave in nombres_visibles and not es_combinada(hoja, fila_actual, 2):
            hoja.cell(row=fila_actual, column=2).value = entradas_usuario[i]
            i += 1
        fila_actual += 1

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida, load_workbook(salida, data_only=True)

def obtener_resumen(hoja):
    datos = []
    for fila in hoja.iter_rows(min_row=4, max_col=10, max_row=100):
        c1 = fila[0].value
        c2 = fila[1].value
        c3 = fila[9].value if len(fila) > 9 else None
        if c1 or c2 or c3:
            datos.append((c1, c2, c3))
    return datos

etiquetas, valores_ref = cargar_referencia()
entradas_usuario = []

with st.form("formulario_datos"):
    st.markdown("### Ingreso de datos")
    for clave in nombres_visibles:
        label = etiquetas.get(clave, clave)
        valor_ref = valores_ref[list(etiquetas.keys()).index(clave)]
        entrada = st.number_input(f"{clave} - {label}", value=float(valor_ref) if isinstance(valor_ref, (int, float)) else 0.0, step=1.0, key=f"input_{clave}")
        entradas_usuario.append(entrada)
    calcular = st.form_submit_button("Calcular Tarifa")

if calcular:
    archivo_excel, wb_final = actualizar_excel_con_datos(entradas_usuario, etiquetas)
    hoja_resumen = wb_final["Resumen de Calculo"]
    datos = obtener_resumen(hoja_resumen)

    st.success("✅ Cálculo realizado. A continuación se muestra la hoja resumen.")
    st.write("### 📄 Hoja: Resumen de Cálculo")

    for fila in datos:
        etiqueta = fila[1] if fila[1] else ""
        valor = fila[2] if fila[2] is not None else ""
        if etiqueta:
            st.markdown(f"**{etiqueta}**: {valor}")

    # Generar Excel con solo la hoja de resumen
    resumen_salida = BytesIO()
    df_resumen = pd.DataFrame([{"Código": f[0], "Concepto": f[1], "Valor": f[2]} for f in datos if f[1]])
    with pd.ExcelWriter(resumen_salida, engine="xlsxwriter") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen de Cálculo", index=False)
    resumen_salida.seek(0)

    st.download_button("📥 Descargar Hoja Resumen en Excel", data=resumen_salida, file_name="Resumen_Calculo_ERSEP.xlsx")
